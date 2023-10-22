import openpyxl
import tkinter as tk
from tkinter import filedialog
import webbrowser
import win32com.client as win32
import os


def apply_wholesale_format(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # Hide columns T to Y
    for col_letter in ["T", "U", "V", "W", "X", "Y"]:
        sheet.column_dimensions[col_letter].hidden = True

    # Scroll columns
    for _ in range(2, 13):
        sheet.sheet_view.selection[0].activeCell = f"C{_}"
        sheet.sheet_view.selection[0].sqref = f"C{_}"

    # Set column widths and attribute descriptions
    col_width = 23.71
    columns_with_attributes = ["AC", "AD", "AE", "AF"]
    attribute_descriptions = [
        "Ack Recipient Addressee attribute description",
        "Ack Recipient Addr Attribute Description",
        "Ack Recipient CSZ attribute description",
        "Ack Salutation attribute description",
    ]
    for col_letter in columns_with_attributes:
        sheet.column_dimensions[col_letter].width = col_width
        attribute_desc = attribute_descriptions.pop(0)
        sheet[f"{col_letter}1"] = attribute_desc
    
    #Add filters to the top row
    sheet.auto_filter.ref = sheet.dimensions
    # Freeze panes
    sheet.freeze_panes = "A2"

    #Remove Values in Column K from K2 onwards
    for row in sheet.iter_rows(min_row=2, max_col=11, max_row=sheet.max_row):
        row[10].value = None 
  # Replace "WHITEMAIL" with "WHITE MAIL" in Column D
    for row in sheet.iter_rows(min_row=2, max_col=4, max_row=sheet.max_row):
        if row[3].value and "WHITEMAIL" in row[3].value:
            row[3].value = row[3].value.replace("WHITEMAIL", "WHITE MAIL")

# Set cell value and formulas
    sheet["AB2"] = "Under $500"
    last_row = sheet.max_row
    sheet[f"AB2"].value = '=IF(C2<=4,"None",IF(AND(C2<=499,C2>=5),"Under $500",IF(C2>=500,"Over $500","NA")))'

    # Autofill the formula from AB2 to the last row
    formula_cell = sheet["AB2"]
    for row in range(2, last_row + 1):
        formula = '=IF(C{}<=4,"None",IF(AND(C{}<=499,C{}>=5),"Under $500",IF(C{}>=500,"Over $500","NA")))'.format(row, row, row, row)
        sheet.cell(row=row, column=formula_cell.column, value=formula)

    # Save the changes
    workbook.save(file_path)
    print("Wholesale formatting applied.")

def apply_retail_format(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Set column widths and attribute descriptions
    col_width = 27
    columns_with_attributes = ["N", "O", "P", "Q"]
    attribute_descriptions = [
        "Ack Recipient Addressee attribute description",
        "Ack Recipient Addr Attribute Description",
        "Ack Recipient CSZ attribute description",
        "Ack Salutation attribute description",
    ]
    for col_letter in columns_with_attributes:
        sheet.column_dimensions[col_letter].width = col_width
        attribute_desc = attribute_descriptions.pop(0)
        sheet[f"{col_letter}1"] = attribute_desc

    # Set cell values and formulas
    sheet["L1"] = "TY"
    sheet["M1"] = "Ref"
    sheet["L2"] = "Under $500"
    last_row = sheet.max_row
    formula_cell = sheet["L2"]
    for row in range(2, last_row + 1):
        formula = '=IF(C{}<=4,"None",IF(AND(C{}<=499,C{}>=5),"Under $500",IF(C{}>=500,"Over $500","NA")))'.format(row, row, row, row)
        sheet.cell(row=row, column=formula_cell.column, value=formula)

    # Set number format for column E
    for cell in sheet["E:E"]:
        cell.number_format = "00"

    # Freeze panes
    sheet.freeze_panes = "A2"

    # Apply AutoFilter
    sheet.auto_filter.ref = sheet.dimensions

    # Save the changes
    workbook.save(file_path)
    print("Retail formatting applied.")

def convert_mt_file(file_path):
        # Create a tkinter root window (hidden)
    root = tk.Tk()
    root.withdraw()

    # Check if the user selected a file
    if file_path:
        # Check if the file exists
        if not os.path.exists(file_path):
            print("File not found.")
        else:
            # Create an instance of Excel
            excel = win32.Dispatch("Excel.Application")
        
            # Create a new workbook
            wb = excel.Workbooks.Add()

            # Use QueryTables to import the text file with custom delimiters
            qt = wb.ActiveSheet.QueryTables.Add(Connection=f"TEXT;{file_path}",
                                            Destination=wb.ActiveSheet.Range("A1"))

            # Manually set the constants for delimiter options
            xlDelimited = 1
            xlTextFormatConsecutiveDelimiter = 1
            xlTextFormatTabDelimited = 1
            xlTextFormatSemicolonDelimited = 1
            xlTextFormatCommaDelimited = 1

            # Set the custom delimiter (in this case, a comma)
            qt.TextFileParseType = xlDelimited
            qt.TextFileConsecutiveDelimiter = False
            qt.TextFileTabDelimiter = False
            qt.TextFileSemicolonDelimiter = False
            qt.TextFileCommaDelimiter = True  # Treat commas as text

            # Refresh the query table to load the data
            qt.Refresh()

            # Make Excel visible (optional)
            excel.Visible = True
    else:
        print("No file selected.")

def open_documentation():
    documentation_url = "https://feedmorewny-my.sharepoint.com/:w:/r/personal/treynolds_feedmorewny_org/_layouts/15/Doc.aspx?sourcedoc=%7B050E299C-02CF-4D9E-B9EA-C09B951670A7%7D&file=Data%20Team%20Training.docx&action=default&mobileredirect=true"
    webbrowser.open(documentation_url)

def on_wholesale_button():
    file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        apply_wholesale_format(file_path)

def on_retail_button():
    file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        apply_retail_format(file_path)

def convert_mt_button():
    file_path = filedialog.askopenfilename()
    if file_path:
        convert_mt_file(file_path)


# Create the GUI window
root = tk.Tk()
root.title("M&T Wholesale & Retail Program")

wholesale_button = tk.Button(root, text="Wholesale Format", command=on_wholesale_button)
wholesale_button.pack()

retail_button = tk.Button(root, text="Retail Format", command=on_retail_button)
retail_button.pack()

documentation_button = tk.Button(root, text="Data Entry Training Guide", command=open_documentation)
documentation_button.pack()

mt_button = tk.Button(root, text="Convert Raw M&T File", command=convert_mt_button)
mt_button.pack()

# Set the window size and position
window_width = 200
window_height =150
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_pos = (screen_width - window_width) // 2
y_pos = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x_pos}+{y_pos}")

# Center the buttons
mt_button.place(relx=0.5, rely=0.15, anchor="center")
wholesale_button.place(relx=0.5, rely=0.4, anchor="center")
retail_button.place(relx=0.5, rely=0.6, anchor="center")
documentation_button.place(relx=0.5, rely=0.85, anchor="center")

root.mainloop()